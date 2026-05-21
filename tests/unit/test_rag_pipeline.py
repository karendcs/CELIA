"""Testes unitários do ExcelRAGPipeline."""
from __future__ import annotations

from pathlib import Path
from unittest.mock import MagicMock

import openpyxl
import pytest

from domain.models import AnswerResult, QueryResult


def _build_pipeline(test_settings, mock_embedder, mock_repository, mock_llm):
    from application.prompts.rag_prompt import RAGPrompt
    from application.services.answer_service import AnswerService
    from application.pipelines.rag_pipeline import ExcelRAGPipeline
    svc = AnswerService(mock_embedder, mock_repository, mock_llm, RAGPrompt(), test_settings)
    return ExcelRAGPipeline(svc, test_settings)


def test_pipeline_processes_all_rows(
    test_settings, sample_xlsx, tmp_path, mock_embedder, mock_repository, mock_llm
):
    pipeline = _build_pipeline(test_settings, mock_embedder, mock_repository, mock_llm)
    out_path = tmp_path / "output.xlsx"

    logs = list(pipeline.run(sample_xlsx, out_path, "Seguranca", "Pergunta", "Resposta"))

    assert out_path.exists(), "Arquivo de saída não foi criado"
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    # Verifica que a coluna Resposta foi preenchida
    answers = [ws.cell(row=r, column=2).value for r in range(2, ws.max_row + 1)]
    assert all(a is not None and a != "" for a in answers)


def test_pipeline_generates_ok_log_line(
    test_settings, sample_xlsx, tmp_path, mock_embedder, mock_repository, mock_llm
):
    pipeline = _build_pipeline(test_settings, mock_embedder, mock_repository, mock_llm)
    out_path = tmp_path / "output.xlsx"
    logs = list(pipeline.run(sample_xlsx, out_path, "Seguranca", "Pergunta", "Resposta"))
    combined = "".join(logs)
    assert "[OK] Baixar planilha" in combined


def test_pipeline_invalid_xlsx_path_logs_error(
    test_settings, tmp_path, mock_embedder, mock_repository, mock_llm
):
    pipeline = _build_pipeline(test_settings, mock_embedder, mock_repository, mock_llm)
    out_path = tmp_path / "output.xlsx"
    logs = list(
        pipeline.run(Path("/nonexistent.xlsx"), out_path, "Sheet1", "Q", "A")
    )
    combined = "".join(logs)
    assert "[ERRO]" in combined


def test_pipeline_skips_empty_questions(
    test_settings, tmp_path, mock_embedder, mock_repository, mock_llm
):
    # Cria xlsx com linha vazia na coluna de perguntas
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Seg"
    ws.append(["Pergunta", "Resposta"])
    ws.append(["", ""])
    ws.append(["Pergunta válida?", ""])
    path = tmp_path / "partial.xlsx"
    wb.save(path)

    out_path = tmp_path / "out.xlsx"
    pipeline = _build_pipeline(test_settings, mock_embedder, mock_repository, mock_llm)
    list(pipeline.run(path, out_path, "Seg", "Pergunta", "Resposta"))

    # LLM só deve ter sido chamado 1 vez (a linha vazia é ignorada)
    assert mock_llm.generate.call_count == 1
